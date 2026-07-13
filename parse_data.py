import openpyxl
import json
import datetime
import os

def parse_excel(file_path):
    wb = openpyxl.load_workbook(file_path, data_only=True)
    data = {
        "sheets": {},
        "tracks": {
            "System": [],
            "Web": [],
            "HPC": []
        },
        "all_participants": []
    }
    
    # List of sheets we want to process
    active_sheets = []
    for sheet_name in wb.sheetnames:
        if sheet_name.startswith("System") or sheet_name.startswith("Web") or sheet_name.startswith("HPC"):
            active_sheets.append(sheet_name)
            
    print("Found active sheets:", active_sheets)
    
    for sheet_name in active_sheets:
        sheet = wb[sheet_name]
        rows = list(sheet.iter_rows(values_only=True))
        if len(rows) < 2:
            continue
            
        # Row 0 may contain meet link and grader names
        row0 = rows[0]
        # Row 1 contains column headers
        headers = rows[1]
        
        meet_link = ""
        if row0 and len(row0) > 0 and row0[0] and str(row0[0]).startswith("Link:"):
            meet_link = str(row0[0]).replace("Link:", "").strip()
            
        # Grader names are at index 9, 16, 23 (in Row 0)
        grader1_name = "Grader 1"
        grader2_name = "Grader 2"
        grader3_name = "Grader 3"
        
        has_grading = len(headers) >= 31 and 'Final Score' in headers
        
        if has_grading:
            if len(row0) > 9 and row0[9] is not None:
                grader1_name = str(row0[9]).strip()
            if len(row0) > 16 and row0[16] is not None:
                grader2_name = str(row0[16]).strip()
            if len(row0) > 23 and row0[23] is not None:
                grader3_name = str(row0[23]).strip()
                
        sheet_participants = []
        
        for r_idx, row in enumerate(rows[2:]):
            # Skip empty rows or rows that do not have a participant name
            if len(row) < 5 or row[4] is None or str(row[4]).strip() == '':
                continue
                
            # Parse columns
            board = str(row[0]).strip() if row[0] is not None else sheet_name
            session = str(row[1]).strip() if row[1] is not None else ""
            
            # Normalize session name: "Chiểu" -> "Chiều"
            if session == "Chiểu":
                session = "Chiều"
                
            # Date formatting
            date_val = row[2]
            if isinstance(date_val, datetime.datetime):
                date_str = date_val.strftime("%d/%m")
            elif date_val is not None:
                date_str = str(date_val).strip()
                if "00:00:00" in date_str:
                    # Clean up timestamp if it is printed as string
                    date_str = date_str.split(" ")[0]
                    # Convert YYYY-MM-DD to DD/MM
                    try:
                        parts = date_str.split("-")
                        date_str = f"{parts[2]}/{parts[1]}"
                    except:
                        pass
            else:
                date_str = ""
                
            time_slot = str(row[3]).strip() if row[3] is not None else ""
            student_name = str(row[4]).strip() if row[4] is not None else ""
            
            # Phone number formatting
            phone_val = row[5]
            if phone_val is not None:
                if isinstance(phone_val, float):
                    phone_str = f"0{int(phone_val)}"
                else:
                    phone_str = str(phone_val).strip()
                    if not phone_str.startswith("0") and phone_str.isdigit():
                        phone_str = "0" + phone_str
            else:
                phone_str = ""
                
            project_code = ""
            if len(row) > 6 and row[6] is not None:
                project_code = str(int(row[6])) if isinstance(row[6], (int, float)) else str(row[6]).strip()
                
            mentor_unit = str(row[7]).strip() if len(row) > 7 and row[7] is not None else ""
            
            # Note column
            note = str(row[8]).strip() if len(row) > 8 and row[8] is not None else ""
            
            is_resigned = (note == "")
            
            participant = {
                "board": board,
                "session": session,
                "date": date_str,
                "timeSlot": time_slot,
                "studentName": student_name,
                "phone": phone_str,
                "projectCode": project_code,
                "mentorUnit": mentor_unit,
                "note": note,
                "isResigned": is_resigned,
                "hasGrading": has_grading
            }
            
            if has_grading:
                # Grader 1
                g1_criteria = [row[c] for c in range(9, 15)]
                g1_total = row[15]
                g1_graded = any(c is not None for c in g1_criteria)
                
                # Grader 2
                g2_criteria = [row[c] for c in range(16, 22)]
                g2_total = row[22]
                g2_graded = any(c is not None for c in g2_criteria)
                
                # Grader 3
                g3_criteria = [row[c] for c in range(23, 29)]
                g3_total = row[29]
                g3_graded = any(c is not None for c in g3_criteria)
                
                # Grader count
                grader_count = sum([1 for g in [g1_graded, g2_graded, g3_graded] if g])
                
                final_score = row[30]
                if isinstance(final_score, (int, float)):
                    final_score = float(final_score)
                else:
                    final_score = 0.0
                    
                participant.update({
                    "graderCount": grader_count,
                    "finalScore": final_score,
                    "graders": [
                        {
                            "name": grader1_name,
                            "criteria": g1_criteria,
                            "total": float(g1_total) if isinstance(g1_total, (int, float)) else 0.0,
                            "graded": g1_graded
                        },
                        {
                            "name": grader2_name,
                            "criteria": g2_criteria,
                            "total": float(g2_total) if isinstance(g2_total, (int, float)) else 0.0,
                            "graded": g2_graded
                        },
                        {
                            "name": grader3_name,
                            "criteria": g3_criteria,
                            "total": float(g3_total) if isinstance(g3_total, (int, float)) else 0.0,
                            "graded": g3_graded
                        }
                    ]
                })
            else:
                participant.update({
                    "graderCount": 0,
                    "finalScore": None,
                    "graders": []
                })
                
            sheet_participants.append(participant)
            data["all_participants"].append(participant)
            
        data["sheets"][sheet_name] = {
            "meetLink": meet_link,
            "participants": sheet_participants
        }
        
        # Categorize by track
        if sheet_name.startswith("System"):
            data["tracks"]["System"].extend(sheet_participants)
        elif sheet_name.startswith("Web"):
            data["tracks"]["Web"].extend(sheet_participants)
        elif sheet_name.startswith("HPC"):
            data["tracks"]["HPC"].extend(sheet_participants)
            
    # Save to data.json
    with open("data.json", "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
        
    print(f"Successfully processed {len(data['all_participants'])} participants and saved to data.json!")

if __name__ == "__main__":
    parse_excel("spreadsheet.xlsx")
